"""Extracts an XLSX sheet's grid (values, formulas, merges, styling) to the
JSON format consumed by docs/assets/calc-sheet.js for live in-browser
recalculation. Requires openpyxl (pip install openpyxl).
"""
import json
import sys
import openpyxl


def extract(path, sheet_name):
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    ws = wb_formulas[sheet_name]
    wsv = wb_values[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    cells = {}
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            r, c = cell.row - 1, cell.column - 1
            key = f"{r},{c}"
            entry = {}
            if isinstance(v, str) and v.startswith("="):
                entry["f"] = v
                entry["cached"] = wsv.cell(row=cell.row, column=cell.column).value
            else:
                entry["v"] = v
            if cell.font and cell.font.bold:
                entry["bold"] = True
            fill = cell.fill
            if (fill and fill.fill_type == "solid" and fill.fgColor
                    and isinstance(fill.fgColor.rgb, str) and fill.fgColor.rgb != "00000000"):
                entry["fill"] = "#" + fill.fgColor.rgb[2:]
            cells[key] = entry

    merges = [
        {"r1": m.min_row - 1, "c1": m.min_col - 1, "r2": m.max_row - 1, "c2": m.max_col - 1}
        for m in ws.merged_cells.ranges
    ]

    return {"maxRow": max_row, "maxCol": max_col, "cells": cells, "merges": merges}


if __name__ == "__main__":
    path, sheet_name, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    data = extract(path, sheet_name)
    with open(out_path, "w") as f:
        json.dump(data, f, indent=0)
    print(f"wrote {out_path}: {len(data['cells'])} cells, {len(data['merges'])} merges")
